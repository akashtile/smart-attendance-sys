from flask import Flask, render_template, request, jsonify
import os, cv2, base64, sqlite3
import numpy as np
from datetime import datetime
from deepface import DeepFace
from sklearn.metrics.pairwise import cosine_similarity
import pandas as pd
from flask import send_file
from openpyxl import Workbook, load_workbook
import pandas as pd
from flask import send_file
from flask import session, redirect, url_for
import os

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "database", "attendance.db")


app = Flask(__name__)
app.secret_key = "supersecretkey"  # can be anything


DATASET = "dataset"
ENC_DIR = "encodings"
DB = "database/attendance.db"

os.makedirs(DATASET, exist_ok=True)
os.makedirs(ENC_DIR, exist_ok=True)
os.makedirs("database", exist_ok=True)

# ---------- DB ----------
def init_db():
    conn = sqlite3.connect(DB)
    cur = conn.cursor()
    cur.execute("""
    CREATE TABLE IF NOT EXISTS attendance (
        id INTEGER PRIMARY KEY,
        name TEXT, roll TEXT, erp TEXT,
        date TEXT, time TEXT, status TEXT
    )
    """)
    conn.commit()
    conn.close()

init_db()
EXCEL_FILE = "attendance.xlsx"
# ---------- Load model ----------
print("Loading model...")
model = DeepFace.build_model("SFace")
print("Model loaded.")

# ---------- Load encodings ----------
def load_encodings():
    try:
        return np.load("encodings/faces.npy"), np.load("encodings/names.npy")
    except:
        return np.array([]), np.array([])

encodings, names = load_encodings()

# ---------- Encode dataset ----------
def rebuild_encodings():
    global encodings, names
    all_enc = []
    all_names = []

    for folder in os.listdir(DATASET):
        path = os.path.join(DATASET, folder)
        for img in os.listdir(path):
            image = cv2.imread(os.path.join(path, img))
            rep = DeepFace.represent(image, model_name="SFace", enforce_detection=False)
            all_enc.append(rep[0]["embedding"])
            all_names.append(folder)

    encodings = np.array(all_enc)
    names = np.array(all_names)

    np.save("encodings/faces.npy", encodings)
    np.save("encodings/names.npy", names)

# ---------- Recognition ----------
def recognize_face(frame):
    if len(encodings) == 0:
        return None

    rep = DeepFace.represent(frame, model_name="SFace", enforce_detection=False)[0]["embedding"]
    sims = cosine_similarity([rep], encodings)[0]
    idx = np.argmax(sims)

    if sims[idx] > 0.45:
        return names[idx]
    return None

def login_required():
    return "user" in session

def admin_required():
    return session.get("role") == "admin"


# ---------- Routes ----------
@app.route("/")
@app.route("/dashboard")
def dashboard():
    if "user" not in session:
        return redirect("/login")

    return render_template("dashboard.html", role=session["role"])


@app.route("/live")
def live():
    return render_template("live.html")

@app.route("/register")
def register():
    return render_template("register.html")

@app.route("/records")
def records():
    return render_template("records.html")

@app.route("/students")
def students():
    return render_template("students.html")

# ---------- API ----------
@app.route("/api/recognize", methods=["POST"])
def recognize():
    global encodings, names

    if len(encodings) == 0:
        return jsonify({"error": "No registered students"}), 400

    data = request.json["image"]
    img = base64.b64decode(data.split(",")[1])
    nparr = np.frombuffer(img, np.uint8)
    frame = cv2.imdecode(nparr, cv2.IMREAD_COLOR)

    result = recognize_face(frame)

    print("Recognition result:", result)  # 👈 DEBUG LINE

    if result:
        try:
            name, roll, erp = result.split("_")
        except:
            return jsonify({"error": "Bad folder format"}), 500

        marked = mark_attendance(name, roll, erp)

        return jsonify({
            "name": name,
            "roll": roll,
            "erp": erp,
            "marked": marked
        })

    return jsonify({"name": "UNKNOWN"})

@app.route("/login")
def login_page():
    return render_template("login.html")

@app.route("/")
def home():
    if "user" in session:
        return redirect("/dashboard")
    return redirect("/login")

@app.route("/student-dashboard")
def student_dashboard():
    return render_template("student_dashboard.html")

@app.route("/api/attendance/filter")
def filter_records():
    q = request.args.get("q", "")
    date = request.args.get("date", "")

    conn = sqlite3.connect(DB)
    cur = conn.cursor()

    query = """
    SELECT name, roll, erp, date, time, status
    FROM attendance
    WHERE (name LIKE ? OR roll LIKE ? OR erp LIKE ?)
    """
    params = [f"%{q}%"] * 3

    if date:
        query += " AND date = ?"
        params.append(date)

    query += " ORDER BY date DESC, time DESC"

    cur.execute(query, params)
    rows = cur.fetchall()
    conn.close()

    return jsonify([
        {"name": r[0], "roll": r[1], "erp": r[2], "date": r[3], "time": r[4], "status": r[5]}
        for r in rows
    ])

@app.route("/api/login", methods=["POST"])
def login():
    data = request.json
    username = data["username"]
    password = data["password"]
    role = data["role"]   # 👈 ADD THIS

    if role == "student":
        return jsonify(success=True, redirect="/student-dashboard")

    if role == "teacher":
        return jsonify(success=True, redirect="/teacher-dashboard")

    return jsonify(success=False)


    conn = sqlite3.connect(DB)
    cur = conn.cursor()
    cur.execute("SELECT role FROM users WHERE username=? AND password=?", (username, password))
    row = cur.fetchone()
    conn.close()

    if row:
        session["user"] = username
        session["role"] = row[0]
        return jsonify({"success": True, "role": row[0]})

    return jsonify({"success": False})

@app.route("/student-records")
def student_records():
    return render_template("student_records.html")

@app.route("/student-notifications")
def student_notifications():
    return render_template("student_notifications.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")

# ---------- TEACHER PAGES ----------

@app.route("/teacher-dashboard")
def teacher_dashboard():
    return render_template("teacher_dashboard.html", role="teacher")

@app.route("/live")
def live_attendance():
    return render_template("live_attendance.html")

@app.route("/register")
def register_student():
    return render_template("register_student.html")

@app.route("/records")
def attendance_records():
    return render_template("attendance_records.html")

@app.route("/students")
def students_list():
    return render_template("students.html")


@app.route("/students")
def students_page():
    return render_template("students.html")

@app.route("/api/students")
def get_students():
    conn = sqlite3.connect(DB)
    cur = conn.cursor()
    cur.execute("SELECT id, name, roll, erp FROM students")
    rows = cur.fetchall()
    conn.close()

    return jsonify([
        {"id": r[0], "name": r[1], "roll": r[2], "erp": r[3]}
        for r in rows
    ])

@app.route("/api/students/delete/<int:id>", methods=["DELETE"])
def delete_student(id):
    conn = sqlite3.connect(DB)
    cur = conn.cursor()
    cur.execute("DELETE FROM students WHERE id=?", (id,))
    conn.commit()
    conn.close()
    return jsonify({"success": True})


def mark_attendance(name, roll, erp):
    conn = sqlite3.connect(DB)
    cur = conn.cursor()

    today = datetime.now().strftime("%Y-%m-%d")
    now = datetime.now().strftime("%H:%M:%S")

    # Prevent duplicate entry for same day
    cur.execute("SELECT * FROM attendance WHERE name=? AND date=?", (name, today))
    if cur.fetchone():
        conn.close()
        return False

    # Insert into database
    cur.execute("""
        INSERT INTO attendance(name, roll, erp, date, time, status)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (name, roll, erp, today, now, "PRESENT"))

    conn.commit()
    conn.close()

    mark_attendance(...)
    export_db_to_excel()


    # ✅ Also update Excel automatically
    write_to_excel(name, roll, erp, today, now, "PRESENT")

    print(f"✅ Saved: {name} in DB + Excel")
    return True

    conn = sqlite3.connect(DB)
    cur = conn.cursor()
    today = datetime.now().strftime("%Y-%m-%d")

    cur.execute("SELECT * FROM attendance WHERE name=? AND date=?", (name, today))
    if cur.fetchone():
        conn.close()
        return False  # already marked

    now = datetime.now()
    cur.execute("""
    INSERT INTO attendance(name, roll, erp, date, time, status)
    VALUES (?, ?, ?, ?, ?, ?)
    """, (name, roll, erp, now.strftime("%Y-%m-%d"), now.strftime("%H:%M:%S"), "PRESENT"))

    conn.commit()
    conn.close()

    print(f"✔ Attendance saved: {name}")
    return True



@app.route("/api/register", methods=["POST"])
def api_register():
    name = request.json["name"]
    roll = request.json["roll"]
    erp = request.json["erp"]

    conn = sqlite3.connect(DB)
    cur = conn.cursor()

    # create students table if not exists
    cur.execute("""
    CREATE TABLE IF NOT EXISTS students (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT,
        roll TEXT UNIQUE,
        erp TEXT UNIQUE,
        registered_date TEXT
    )
    """)

    # insert student
    cur.execute("""
        INSERT OR IGNORE INTO students(name, roll, erp, registered_date)
        VALUES (?, ?, ?, ?)
    """, (name, roll, erp, datetime.now().strftime("%Y-%m-%d")))

    conn.commit()
    conn.close()

    folder = f"{name}_{roll}_{erp}"
    os.makedirs(os.path.join(DATASET, folder), exist_ok=True)

    return jsonify({"folder": folder})


@app.route("/api/export/excel")
def export_excel():
    conn = sqlite3.connect(DB)

    # Read attendance table into pandas
    df = pd.read_sql_query("""
        SELECT name, roll, erp, date, time, status
        FROM attendance
        ORDER BY date DESC, time DESC
    """, conn)

    conn.close()

    # Save to Excel
    filename = "attendance_report.xlsx"
    df.to_excel(filename, index=False)

    return send_file(filename, as_attachment=True)

@app.route("/api/students/update/<int:id>", methods=["POST"])
def update_student(id):
    data = request.json
    conn = sqlite3.connect(DB)
    cur = conn.cursor()

    cur.execute("""
        UPDATE students SET name=?, roll=?, erp=?
        WHERE id=?
    """, (data["name"], data["roll"], data["erp"], id))

    conn.commit()
    conn.close()
    return jsonify({"success": True})


@app.route("/api/save-images", methods=["POST"])
def save_images():
    folder = request.json["folder"]
    images = request.json["images"]

    path = os.path.join(DATASET, folder)

    if not os.path.exists(path):
        return jsonify({"error": "Folder does not exist"}), 400

    count = 0

    for i, img in enumerate(images):
        data = base64.b64decode(img.split(",")[1])
        arr = np.frombuffer(data, np.uint8)
        img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
        cv2.imwrite(f"{path}/{i}.jpg", img)
        count += 1

    rebuild_encodings()

    return jsonify({"success": True, "saved": count})

@app.route("/api/stats")
def api_stats():
    conn = sqlite3.connect(DB)
    cur = conn.cursor()

    today = datetime.now().strftime("%Y-%m-%d")

    # total students
    cur.execute("SELECT COUNT(*) FROM students")
    total = cur.fetchone()[0]

    # present today
    cur.execute("SELECT COUNT(DISTINCT name) FROM attendance WHERE date=?", (today,))
    present = cur.fetchone()[0]

    absent = total - present if total >= present else 0

    conn.close()

    print("STATS:", total, present, absent)  # DEBUG

    return jsonify({
        "total": total,
        "present": present,
        "absent": absent
    })

def export_db_to_excel():
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    cur.execute("SELECT date, time, name, roll, erp, status FROM attendance")
    rows = cur.fetchall()

    wb = Workbook()
    sheet = wb.active
    sheet.append(["Date", "Time", "Name", "Roll", "ERP", "Status"])

    for row in rows:
        sheet.append(row)

    wb.save("attendance.xlsx")
    conn.close()



def write_to_excel(name, roll, erp, date, time, status):
    # Create file with headers if not exists
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Roll", "ERP", "Date", "Time", "Status"])
        wb.save(EXCEL_FILE)

    # Append new row
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([name, roll, erp, date, time, status])
    wb.save(EXCEL_FILE)

@app.route("/api/attendance/all")
def all_attendance():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    cur.execute("""
        SELECT name, roll, erp, date, time, status
        FROM attendance
        ORDER BY date DESC, time DESC
    """)

    rows = cur.fetchall()
    conn.close()

    return jsonify([dict(r) for r in rows])

@app.route("/api/export/excel")
def export_excel_report():
    conn = sqlite3.connect(DB)

    df = pd.read_sql_query("""
        SELECT name, roll, erp, date, time, status
        FROM attendance
        ORDER BY date DESC, time DESC
    """, conn)

    conn.close()

    filename = "attendance_report.xlsx"
    df.to_excel(filename, index=False)

    return send_file(filename, as_attachment=True)


if __name__ == "__main__":
    app.run(debug=True)
